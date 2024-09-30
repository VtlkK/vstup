import openpyxl
from django.http import HttpResponse
from openpyxl.styles import PatternFill, Border, Side
from .models import Sing_Clients
from io import BytesIO

def exel_data(request):
    data = Sing_Clients.objects.all()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data'

    headers = ['id',  'Прізвище', "Ім'я", 'По батькові', 'Стать', 'Дата народження', 'Номер телефону', 'Область']
    worksheet.append(headers)
    worksheet.row_dimensions[1].height = 20
    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 45

    for item in data:
        worksheet.append([item.id ,item.soname, item.name, item.fname, item.sex, item.date, item.num, item.born])
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    redfill = PatternFill(start_color='0000FFFF', end_color='FF0000', fill_type='solid')
    worksheet['A1'].fill = redfill
    worksheet['A1'].border = thin_border
    worksheet['B1'].fill = redfill
    worksheet['B1'].border = thin_border
    worksheet['C1'].fill = redfill
    worksheet['C1'].border = thin_border
    worksheet['D1'].fill = redfill
    worksheet['D1'].border = thin_border
    worksheet['E1'].fill = redfill
    worksheet['E1'].border = thin_border
    worksheet['F1'].fill = redfill
    worksheet['F1'].border = thin_border
    worksheet['G1'].fill = redfill
    worksheet['G1'].border = thin_border
    worksheet['H1'].fill = redfill
    worksheet['H1'].border = thin_border
    worksheet['I1'].fill = redfill
    worksheet['I1'].border = thin_border

    worksheet['M1'] = 'Кількість осіб'
    worksheet['M2'] = '=COUNTA(A2:A40)'
    worksheet.column_dimensions['M'].width = 15

    yellowfill = PatternFill(start_color='00FFFF99', end_color='00FFFF99', fill_type='solid')
    worksheet['M2'].fill = yellowfill

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=db_exel.xlsx'

    return response



